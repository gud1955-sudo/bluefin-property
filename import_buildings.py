import sys, os, re
sys.stdout.reconfigure(encoding='utf-8')
os.chdir('E:/bluefin-property')
sys.path.insert(0, 'E:/bluefin-property')

import openpyxl
from app import app, db, Building, Contact

PHONE_RE = re.compile(r'^[\d\s\-/+().]{7,}$')

def is_phone(text):
    if not text:
        return False
    t = str(text).strip()
    return bool(PHONE_RE.match(t)) and sum(c.isdigit() for c in t) >= 6

def split_contact(text):
    t = str(text).strip()
    if ')' in t:
        parts = t.split(')', 1)
        return parts[0].strip(), parts[1].strip()
    return '', t

with app.app_context():
    wb = openpyxl.load_workbook('전화부호부.xlsx')

    building_map = {}  # (addr, name) -> building_id
    total_buildings = 0
    total_contacts = 0

    def get_or_create_building(addr, bname):
        global total_buildings
        key = (addr.strip(), bname.strip())
        if key not in building_map:
            b = Building(name=bname.strip(), address=addr.strip())
            db.session.add(b)
            db.session.flush()
            building_map[key] = b.id
            total_buildings += 1
        return building_map[key]

    def add_contact(bid, person, phone):
        global total_contacts
        if not person:
            return
        title, name = split_contact(str(person).strip())
        c = Contact(
            building_id=bid,
            name=name or str(person).strip(),
            phone=str(phone).strip() if phone else '',
            title=title
        )
        db.session.add(c)
        total_contacts += 1

    # ── 숙주 시트
    ws = wb['숙주']
    for row in ws.iter_rows(min_row=2, values_only=True):
        addr, bname, person, phone = row[0], row[1], row[2], row[3]
        if not bname:
            continue
        bid = get_or_create_building(addr or '', bname)
        if person:
            add_contact(bid, person, phone)

    # ── 임대 시트
    ws2 = wb['임대']
    vals = [row[0] for row in ws2.iter_rows(min_row=4, values_only=True)]

    i = 0
    n = len(vals)
    while i < n:
        v = vals[i]
        if v is None or str(v).strip() in ('', '임대 건물'):
            i += 1
            continue

        addr = str(v).strip()
        i += 1
        if i >= n or vals[i] is None:
            continue

        bname = str(vals[i]).strip()
        i += 1

        bid = get_or_create_building(addr, bname)

        # contacts until empty row
        while i < n and vals[i] is not None:
            line = str(vals[i]).strip()
            i += 1
            if not line or line == '(비어 있음)':
                continue
            phone = ''
            if i < n and vals[i] is not None and is_phone(str(vals[i])):
                phone = str(vals[i]).strip()
                i += 1
            add_contact(bid, line, phone)

    db.session.commit()
    print(f'완료: 건물 {total_buildings}개, 연락처 {total_contacts}개 등록')
